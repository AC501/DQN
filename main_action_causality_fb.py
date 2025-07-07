from openpyxl.reader.excel import load_workbook

from environment.sumotl_action_causality_yuanshi import SUMOTrafficLights
from learner.q_learning import QLearner
from exploration.epsilon_greedy import EpsilonGreedy
import datetime
import warnings
from replay import ReplayBuffer
import os
import torch
import pandas as pd
from torch.utils.data import DataLoader, Dataset
import torch.nn as nn
import numpy as np
import torch.optim as optim
from RewardFunction.FunctionFitting_fuben3 import rewardfitmain
from CPM.cpmDataStructure import process_cpm_data


warnings.filterwarnings(action='ignore')




# 将列解析为数值数组
def parse_state_column(column):
    return [list(map(float, row.split(', '))) for row in column]




# 滑动窗口处理输入数据
def create_sliding_window(states_rewards, y, window_size):
    inputs = []
    conditions = []
    for i in range(len(states_rewards) - window_size + 1):
        inputs.append(states_rewards[i:i + window_size])  # 滑动窗口
        sum_y = np.sum(y[i:i + window_size])  # 计算每个窗口内y值的总和
        conditions.append([sum_y] * window_size)
    return np.array(inputs), np.array(conditions)

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")



# 数据集和 DataLoader
class TrafficDataset(Dataset):
    def __init__(self, states_rewards, conditions):
        self.states = states_rewards
        self.conditions = conditions

    def __len__(self):
        return len(self.states)

    def __getitem__(self, idx):
        return self.states[idx], self.conditions[idx]




# 定义前馈神经网络作为扩散模型
class DiffusionModel(nn.Module):
    def __init__(self, input_dim, condition_dim, hidden_dim=128):
        super(DiffusionModel, self).__init__()
        self.fc1 = nn.Linear(input_dim + condition_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, input_dim)

    def forward(self, x, condition):
        # condition = condition.unsqueeze(-1) # [batch_size, seq_len, 1]
        print('x',x.shape)
        print('con',condition.shape)
        x = torch.cat([x, condition], dim=2)
        x = torch.relu(self.fc1(x))

        x = torch.relu(self.fc2(x))
        x = self.fc3(x)
        return x


# 定义扩散过程中的噪声函数
def add_noise(x, sigma,device):
    noise = torch.randn_like(x).to(device)
    return x + sigma * noise


# 定义训练函数
def train_diffusion_model(model, dataloader, optimizer, num_epochs, device):
    best_model_path = "D:/tsc_ddqn_prb_1Con_new/CPM/BestDiffusionModel.pth"
    # 如果存在预训练模型，则加载
    if os.path.exists("DiffusionModel111.pth"):
        model.load_state_dict(torch.load("DiffusionModel111.pth"))
    model.train()  # 设置为训练模式
    best_loss = float('inf')  # 初始化最小损失值为正无穷
    for epoch in range(num_epochs):
        epoch_loss = 0  # 记录每个epoch的总损失
        for batch_x, batch_c in dataloader:
            batch_x, batch_c = batch_x.to(device), batch_c.to(device)
            # 确保条件变量的形状正确
            batch_c = batch_c.unsqueeze(-1)
            # 清除之前的梯度
            optimizer.zero_grad()
            # 添加噪声
            noisy_x = add_noise(batch_x, sigma=0.1, device=device)
            # 前向传播
            predicted_x = model(noisy_x, batch_c)
            # 计算损失（MSE）
            loss = torch.mean((predicted_x - batch_x) ** 2)
            epoch_loss += loss.item()  # 累加当前batch的损失
            # 反向传播
            loss.backward()
            # 更新模型参数
            optimizer.step()

        # # 输出当前epoch的平均损失
        # if (epoch + 1) % 10 == 0:
        #     print(f"Epoch [{epoch + 1}/{num_epochs}], Loss: {epoch_loss / len(dataloader):.4f}")
        # 如果当前epoch的损失小于最优损失，则保存模型
        if epoch_loss < best_loss:
            best_loss = epoch_loss
            torch.save(model.state_dict(), best_model_path)  # 保存最优模型
            print(f"模型参数已保存，当前最佳损失：{best_loss:.4f}")


# 定义生成新样本的函数
def generate_new_samples(model, condition, num_samples, device):
    if os.path.exists("BestDiffusionModel.pth"):
        model.load_state_dict(torch.load("BestDiffusionModel.pth"))
    model.eval()
    z = torch.randn(num_samples, 5, model.fc1.in_features - condition.shape[1]).to(device)
    condition_expanded = condition.repeat(num_samples, 5).unsqueeze(-1)
    # input_data = torch.cat([z, condition_expanded], dim=1)

    with torch.no_grad():
        generated_samples = model(z, condition_expanded)
    return generated_samples

def main():
    global generated_samples
    epoches = 3
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    for epoch in range(epoches):

        # 创建SUMO环境
        env = SUMOTrafficLights('nets/22grid_fuben_e1.sumocfg', 8813, True, 32)
        # 创建经验回放缓冲区
        learners = {}
        replay_buffers = {}
        # 为每个交通信号灯创建学习者
        for tlID in env.get_trafficlights_ID_list():
            replay_buffers[tlID] = ReplayBuffer(capacity=10000)
            learners[tlID] = QLearner(tlID, env, 0, 0, 0.01, 0.01, 'train', replay_buffers[tlID], 15, 1.0, 0.1, 0.001,
                                      5,
                                      2, 32, "weights/best_model.pth")
        # 设置 learners 和 replay_buffers
        env.learners = learners
        env.replay_buffers = replay_buffers

        # 设置回合数
        n_episodes = 1
        # 运行回合
        for i in range(n_episodes):
            print("Episode =", i, "====================")
            arq_avg_nome = 'tl_%d.txt' % i
            arq_tl = open(arq_avg_nome, 'w')
            arq_tl.writelines('##%s## \n' % datetime.datetime.now().time())

            env.run_episode(5000, None)

            arq_tl.close()

        #训练条件扩散模型和拟合奖励函数
        input_filepath = 'D:/tsc_ddqn_prb_1Con_new/data/CPMData.xlsx'
        output_filepath = 'D:/tsc_ddqn_prb_1Con_new/data/CPMData_final.xlsx'
        process_cpm_data(input_filepath, output_filepath)
        # 读取state和reward数据
        SRData = pd.read_excel(r'D:\tsc_ddqn_prb_xiugaimainkuangjia\data\CPMData_final.xlsx')

        # 读取对应时间内区域产生的车辆数
        ProductData = pd.read_excel(
            r'D:\tsc_ddqn_prb_xiugaimainkuangjia\data\produced_vehicles.xlsx')
        # 读取对应时间内区域离开的车辆数
        LeaveData = pd.read_excel(r'D:\tsc_ddqn_prb_xiugaimainkuangjia\data\departed_vehicles.xlsx')
        print(f"SRData:{len(SRData)},ProductData:{len(ProductData)},LeaveData:{len(LeaveData)}")
        # 向SRData数据集添加y
        SRData['y'] = ProductData['Produced Vehicles'] - LeaveData['Departed Vehicles']

        states_rewards = []
        for col in ['state_reward_0', 'state_reward_1', 'state_reward_2', 'state_reward_3']:
            states_rewards.append(parse_state_column(SRData[col]))

        # 合并所有状态奖励列
        states_rewards = np.concatenate(states_rewards, axis=1)  # 合并为一个大的状态矩阵
        y = SRData['y'].values

        window_size = 5
        input_states_rewards, input_y = create_sliding_window(states_rewards, y, window_size)
        #将得到的x和y保存为excel



        # 转换为 Tensor
        states_rewards_tensor = torch.tensor(input_states_rewards, dtype=torch.float32).to(device)
        y_tensor = torch.tensor(input_y, dtype=torch.float32).to(device)

        # 超参数定义
        input_dim = input_states_rewards.shape[2]
        condition_dim = 1
        hidden_dim = 128
        num_epochs = 300
        batch_size = 32
        learning_rate = 0.0001
        num_samples = 20  # 生成样本的数量
        # 创建模型、定义优化器
        model = DiffusionModel(input_dim, condition_dim, hidden_dim).to(device)
        optimizer = optim.Adam(model.parameters(), lr=learning_rate)

        # 初始数据集
        dataset_x = states_rewards_tensor.to(device)  # Ensure this is on the same device
        dataset_y = y_tensor.unsqueeze(-1).to(device)  # Ensure this is on the same device

        # 将数据集包装成DataLoader
        dataset = TrafficDataset(states_rewards_tensor, y_tensor)
        data_loader = DataLoader(dataset, batch_size=32, shuffle=True)

        # 预训练扩散模型
        train_diffusion_model(model, data_loader, optimizer, num_epochs, device)

        # 条件
        new_condition = torch.tensor([[0]]).to(device)

        # 将 datasetD 转移到同一设备
        datasetD = torch.cat((dataset_x, dataset_y), dim=-1)
        # 开始迭代优化
        for i in range(10):
            # 生成符合条件的新样本 x
            generated_samples = generate_new_samples(model, new_condition, num_samples, device)

        # 假设 generated_samples 是一个形状为 (50, 5, 24) 的 ndarray 或 tensor
        generated_samples_cpu = generated_samples.cpu().numpy()  # 如果是 tensor，要先转换到 CPU

        # 我们需要将每列的 24 个特征分为 4 个 state_reward，每个 state_reward 包含 6 个数值
        reshaped_samples = generated_samples_cpu.reshape(-1, 4, 6)  # 将每行分成 4 个 state_reward，每个 state_reward 包含 6 个数值
        # 保留整数值，确保数据是整数而非浮动数
        reshaped_samples = np.round(reshaped_samples).astype(int)  # 转换为整数类型

        # 将每个 state_reward_X 组按要求拼接成字符串
        merged_samples = []
        for i in range(5 * num_samples):
            state_rewards = []
            for j in range(4):  # 对于每行的 4 个 state_reward 列
                state_rewards.append(', '.join(map(str, reshaped_samples[i, j])))  # 将每个 state_reward 转为逗号分隔的字符串
            merged_samples.append(state_rewards)

        # 转换为 DataFrame，列名为 state_reward_0, state_reward_1, ..., state_reward_3
        df = pd.DataFrame(merged_samples, columns=[f"state_reward_{i}" for i in range(4)])
        # Excel 文件路径
        file_path = 'D:\\tsc_ddqn_prb_1Con_new\\data\\CPMData_final.xlsx'
        # 使用 openpyxl 加载现有的 Excel 文件
        book = load_workbook(file_path)
        # 如果 Sheet1 已经存在，加载该工作表；如果不存在，可以创建一个新的工作表
        if 'Sheet1' in book.sheetnames:
            sheet = book['Sheet1']
        else:
            sheet = book.create_sheet('Sheet1')
        # 获取追加数据之前的行数
        start_row = sheet.max_row + 1

        # 将新的 DataFrame 数据追加到现有工作表的末尾
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                # 将数据写入到指定单元格
                sheet.cell(row=start_row + i, column=j + 1, value=value)
        # 保存文件
        book.save(file_path)

        # 调用 FunctionFitting.py 进行状态奖励数据处理和模型训练
        rewardfitmain()

if __name__ == '__main__':
    main()
