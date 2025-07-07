import os
import pandas as pd
from torch.utils.data import DataLoader, Dataset
import torch
import torch.nn as nn
import torch.optim as optim
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


#读取state和reward数据
SRData = pd.read_excel(r'D:\tsc_ddqn_prb\data\CPMData_final.xlsx')

#读取对应时间内区域产生的车辆数
ProductData = pd.read_excel(r'D:\tsc_ddqn_prb\data\produced_vehicles.xlsx')

#读取对应时间内区域离开的车辆数
LeaveData = pd.read_excel(r'D:\tsc_ddqn_prb\data\departed_vehicles.xlsx')

print(f"SRdataLength:{len(SRData)},ProductLen:{len(ProductData)},LeaveLen:{len(LeaveData)}")

#向SRData数据集添加y
SRData['y'] = ProductData['Produced Vehicles'] - LeaveData['Departed Vehicles']
# 将列解析为数值数组
def parse_state_column(column):
    return [list(map(float, row.split(', '))) for row in column]

states_rewards = []
for col in ['state_reward_0', 'state_reward_1', 'state_reward_2', 'state_reward_3']:
    states_rewards.append(parse_state_column(SRData[col]))

# 合并所有状态奖励列
states_rewards = np.concatenate(states_rewards, axis=1)  # 合并为一个大的状态矩阵
y = SRData['y'].values


# 滑动窗口处理输入数据
def create_sliding_window(states_rewards, y, window_size):
    inputs = []
    conditions = []
    for i in range(len(states_rewards) - window_size + 1):
        inputs.append(states_rewards[i:i + window_size])  # 滑动窗口
        sum_y = np.sum(y[i:i + window_size])  # 计算每个窗口内y值的总和
        conditions.append([sum_y] * window_size)
    return np.array(inputs), np.array(conditions)

window_size = 5
input_states_rewards, input_y = create_sliding_window(states_rewards, y, window_size)
# 转换为 Tensor
states_rewards_tensor = torch.tensor(input_states_rewards, dtype=torch.float32)
y_tensor = torch.tensor(input_y, dtype=torch.float32)



# 数据集和 DataLoader
class TrafficDataset(Dataset):
    def __init__(self, states_rewards, conditions):
        self.states = states_rewards
        self.conditions = conditions

    def __len__(self):
        return len(self.states)

    def __getitem__(self, idx):
        return self.states[idx], self.conditions[idx]


dataset = TrafficDataset(states_rewards_tensor, y_tensor)
data_loader = DataLoader(dataset, batch_size=32, shuffle=True)



# 定义一个简单的前馈神经网络作为扩散模型
class DiffusionModel(nn.Module):
    def __init__(self, input_dim, condition_dim, hidden_dim=128):
        super(DiffusionModel, self).__init__()
        self.fc1 = nn.Linear(input_dim + condition_dim, hidden_dim)
        self.fc2 = nn.Linear(hidden_dim, hidden_dim)
        self.fc3 = nn.Linear(hidden_dim, input_dim)
        # 条件嵌入层
        self.condition_embed = nn.Embedding(condition_dim, hidden_dim)

    def forward(self, x, condition):
        # condition = condition.unsqueeze(-1) # [batch_size, seq_len, 1]
        x = torch.cat([x, condition], dim=2)
        x = torch.relu(self.fc1(x))
        x =torch.relu(self.fc2(x))
        x = self.fc3(x)
        return x


# 定义扩散过程中的噪声函数
def add_noise(x, sigma, device):
    noise = torch.randn_like(x).to(device)
    return x + sigma * noise




# 定义训练函数
def train_diffusion_model(model, dataloader, optimizer, num_epochs, device):
    best_model_path = "D:/tsc_ddqn_prb_1Con_new/CPM/BestDiffusionModel.pth"
    # 如果存在预训练模型，则加载
    if os.path.exists("DiffusionModel111.pth"):
        model.load_state_dict(torch.load("DiffusionModel111.pth"))
    model.train()  # 设置为训练模式
    best_loss = float('inf')  # 初始化最小损失值为正无穷
    for epoch in range(num_epochs):
        epoch_loss = 0  # 记录每个epoch的总损失
        for batch_x, batch_c in dataloader:
            batch_x, batch_c = batch_x.to(device), batch_c.to(device)
            # 确保条件变量的形状正确
            batch_c = batch_c.unsqueeze(-1)
            # 清除之前的梯度
            optimizer.zero_grad()
            # 添加噪声
            noisy_x = add_noise(batch_x, sigma=0.1, device=device)
            # 前向传播
            predicted_x = model(noisy_x, batch_c)
            # 计算损失（MSE）
            loss = torch.mean((predicted_x - batch_x) ** 2)
            epoch_loss += loss.item()  # 累加当前batch的损失
            # 反向传播
            loss.backward()
            # 更新模型参数
            optimizer.step()

        # # 输出当前epoch的平均损失
        # if (epoch + 1) % 10 == 0:
        #     print(f"Epoch [{epoch + 1}/{num_epochs}], Loss: {epoch_loss / len(dataloader):.4f}")
        # 如果当前epoch的损失小于最优损失，则保存模型
        if epoch_loss < best_loss:
            best_loss = epoch_loss
            torch.save(model.state_dict(), best_model_path)  # 保存最优模型
            print(f"模型参数已保存，当前最佳损失：{best_loss:.4f}")

# 定义生成新样本的函数
def generate_new_samples(model, condition, num_samples, device):
    if os.path.exists("BestDiffusionModel.pth"):
        model.load_state_dict(torch.load("BestDiffusionModel.pth"))
    model.eval()
    z = torch.randn(num_samples, 5,model.fc1.in_features - condition.shape[1]).to(device)
    print('z.shape',z.shape)
    condition_expanded = condition.repeat(num_samples, 5).unsqueeze(-1)
    print('condition_expanded.shape',condition_expanded.shape)
    # input_data = torch.cat([z, condition_expanded], dim=1)

    with torch.no_grad():
        generated_samples = model(z, condition_expanded)
        print('generated_samples',generated_samples.shape)
    return generated_samples

def black_box(tensor):
    x1, x2, x3 = tensor[:, 0], tensor[:, 1], tensor[:, 2]
    return x1 * 5 + x2 * 8 + x3 ** 3 - 9 * x1 * x2

def main():
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    # 超参数定义
    input_dim = input_states_rewards.shape[2]
    condition_dim = 1
    hidden_dim = 128
    num_epochs = 10
    batch_size = 32
    learning_rate = 0.001
    num_samples = 20  # 生成样本的数量
    result = []

    # 创建模型、定义优化器
    model = DiffusionModel(input_dim, condition_dim, hidden_dim).to(device)
    optimizer = optim.Adam(model.parameters(), lr=learning_rate)

    # 初始数据集
    dataset_x = states_rewards_tensor.to(device)  # Ensure this is on the same device
    dataset_y = y_tensor.unsqueeze(-1).to(device)  # Ensure this is on the same device
    print('dataset_x', dataset_x.shape)
    print('dataset_y', dataset_y.shape)

    # 将数据集包装成DataLoader
    dataset = TrafficDataset(states_rewards_tensor, y_tensor)
    data_loader = DataLoader(dataset, batch_size=32, shuffle=True)

    # 预训练扩散模型
    train_diffusion_model(model, data_loader, optimizer, num_epochs, device)

    # 条件
    new_condition = torch.tensor([[0]]).to(device)

    # 将 datasetD 转移到同一设备
    datasetD = torch.cat((dataset_x, dataset_y), dim=-1)

    # 开始迭代优化
    for i in range(5):
        # 生成符合条件的新样本 x
        generated_samples = generate_new_samples(model, new_condition, num_samples, device)

        # 假设 generated_samples 是一个形状为 (50, 5, 24) 的 ndarray 或 tensor
        generated_samples_cpu = generated_samples.cpu().numpy()  # 如果是 tensor，要先转换到 CPU

        # 我们需要将每列的 24 个特征分为 4 个 state_reward，每个 state_reward 包含 6 个数值
        reshaped_samples = generated_samples_cpu.reshape(-1, 4, 6)  # 将每行分成 4 个 state_reward，每个 state_reward 包含 6 个数值
        # 保留整数值，确保数据是整数而非浮动数
        reshaped_samples = np.round(reshaped_samples).astype(int)  # 转换为整数类型

        # 将每个 state_reward_X 组按要求拼接成字符串
        merged_samples = []
        for i in range(250):
            state_rewards = []
            for j in range(4):  # 对于每行的 4 个 state_reward 列
                state_rewards.append(', '.join(map(str, reshaped_samples[i, j])))  # 将每个 state_reward 转为逗号分隔的字符串
            merged_samples.append(state_rewards)

        # 转换为 DataFrame，列名为 state_reward_0, state_reward_1, ..., state_reward_3
        df = pd.DataFrame(merged_samples, columns=[f"state_reward_{i}" for i in range(4)])
        # Excel 文件路径
        file_path = 'D:\\tsc_ddqn_prb_1Con_new\\data\\CPMData_final.xlsx'

        # 使用 openpyxl 加载现有的 Excel 文件
        book = load_workbook(file_path)

        # 如果 Sheet1 已经存在，加载该工作表；如果不存在，可以创建一个新的工作表
        if 'Sheet1' in book.sheetnames:
            sheet = book['Sheet1']
        else:
            sheet = book.create_sheet('Sheet1')

        # 获取现有数据的行数
        start_row = sheet.max_row + 1

        # 将新的 DataFrame 数据追加到现有工作表的末尾
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                # 将数据写入到指定单元格
                sheet.cell(row=start_row + i, column=j + 1, value=value)

        # 保存文件
        book.save(file_path)

        print(f"数据已成功追加到 {file_path} 的 'Sheet1' 中！")

        # Ensure generated_samples is on the same device as datasetD
        generated_samples = generated_samples.to(device)  # Ensure it's on the same device

        # # 输入x，通过黑盒，得到y
        y = black_box(generated_samples).unsqueeze(-1).to(device)

        # 调整 y 的形状为 [50, 5, 1]
        y = y.transpose(1, 2)  # 将 y 的形状从 [50, 24, 1] 调整为 [50, 1, 24]
        y = y.expand(-1, 5, -1)  # 将 y 的形状调整为 [50, 5, 24]

        # 如y 压缩到 [50, 5, 1] 的形状，可以执行以下操作
        y = y[:, :, :1]  # 选择每个样本的前 1 列，调整为 [50, 5, 1]

        # Ensure y is on the same device as generated_samples
        y = y.to(device)  # Make sure y is on the same device as generated_samples

        # 将 x 和 y 合并
        new_line = torch.cat((generated_samples, y), dim=2)

        # 扩展进入数据集
        datasetD = torch.cat((datasetD, new_line), dim=0)

        dataset2 = TrafficDataset(datasetD[:, : ,:24], datasetD[:, : ,24])
        dataloader888 = DataLoader(dataset2, batch_size=batch_size, shuffle=True)
        # 训练扩散模型
        train_diffusion_model(model, dataloader888, optimizer, num_epochs, device)


if __name__ == "__main__":
    if os.path.exists('BestDiffusionModel.pth'):
        os.remove('BestDiffusionModel.pth')
    main()
